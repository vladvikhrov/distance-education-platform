import pandas as pd
from datetime import datetime

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("⚠️ Для красивого оформления установите: pip install openpyxl")


def create_students_example():
    """Создает пример файла с учениками"""
    
    students_data = {
        'Логин': [
            'ivanov_a',
            'petrova_m',
            'sidorov_d',
            'kozlova_e',
            'smirnov_i',
            'lebedeva_a',
            'morozov_p',
            'volkova_o',
            'sokolov_n',
            'novikova_t'
        ],
        'Имя': [
            'Александр',
            'Мария',
            'Дмитрий',
            'Елена',
            'Иван',
            'Анастасия',
            'Петр',
            'Ольга',
            'Николай',
            'Татьяна'
        ],
        'Фамилия': [
            'Иванов',
            'Петрова',
            'Сидоров',
            'Козлова',
            'Смирнов',
            'Лебедева',
            'Морозов',
            'Волкова',
            'Соколов',
            'Новикова'
        ],
        'Отчество': [
            'Петрович',
            'Ивановна',
            'Александрович',
            'Сергеевна',
            '',  # Пример без отчества
            'Викторовна',
            'Михайлович',
            'Дмитриевна',
            'Андреевич',
            'Николаевна'
        ],
        'Пароль': [
            'pass1234',
            'maria2024',
            'dima2024',
            'elena123',
            'ivan2024',
            'nastya123',
            'petr2024',
            'olga123',
            'nikolay24',
            'tanya2024'
        ]
    }
    
    df = pd.DataFrame(students_data)
    filename = 'пример_ученики.xlsx'
    df.to_excel(filename, index=False, sheet_name='Ученики')
    
    # Форматирование (если доступен openpyxl)
    if HAS_OPENPYXL:
        format_excel(filename, 'Ученики')
    
    print(f"✅ Создан файл: {filename}")
    return filename


def create_teachers_example():
    """Создает пример файла с учителями"""
    
    teachers_data = {
        'Логин': [
            'teacher_math',
            'teacher_rus',
            'teacher_phys',
            'teacher_chem',
            'teacher_bio',
            'teacher_hist',
            'teacher_eng',
            'teacher_inf'
        ],
        'Имя': [
            'Ольга',
            'Сергей',
            'Анна',
            'Владимир',
            'Екатерина',
            'Михаил',
            'Елена',
            'Дмитрий'
        ],
        'Фамилия': [
            'Николаева',
            'Белов',
            'Соколова',
            'Кузнецов',
            'Павлова',
            'Федоров',
            'Васильева',
            'Орлов'
        ],
        'Отчество': [
            'Викторовна',
            'Михайлович',
            'Петровна',
            'Игоревич',
            'Александровна',
            'Сергеевич',
            'Владимировна',
            'Николаевич'
        ],
        'Пароль': [
            'teacher123',
            'sergey2024',
            'anna_phys',
            'vlad_chem',
            'kate_bio',
            'misha_hist',
            'elena_eng',
            'dima_inf'
        ]
    }
    
    df = pd.DataFrame(teachers_data)
    filename = 'пример_учителя.xlsx'
    df.to_excel(filename, index=False, sheet_name='Учителя')
    
    if HAS_OPENPYXL:
        format_excel(filename, 'Учителя')
    
    print(f"✅ Создан файл: {filename}")
    return filename


def create_classes_subjects_example():
    """Создает пример файла с классами и предметами"""
    
    filename = 'пример_классы_и_предметы.xlsx'
    
    # Создаем Excel writer
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        
        # Класс 9А - основные предметы
        subjects_9a = {
            'Математика': [''] * 5,
            'Русский язык': [''] * 5,
            'Литература': [''] * 5,
            'Физика': [''] * 5,
            'Химия': [''] * 5,
            'Биология': [''] * 5,
            'История': [''] * 5,
            'География': [''] * 5,
            'Английский язык': [''] * 5,
            'Информатика': [''] * 5,
            'Обществознание': [''] * 5,
            'ОБЖ': [''] * 5
        }
        df_9a = pd.DataFrame(subjects_9a)
        df_9a.to_excel(writer, sheet_name='9А', index=False)
        
        # Класс 9Б
        subjects_9b = {
            'Математика': [''] * 5,
            'Русский язык': [''] * 5,
            'Литература': [''] * 5,
            'Физика': [''] * 5,
            'Химия': [''] * 5,
            'Биология': [''] * 5,
            'История': [''] * 5,
            'География': [''] * 5,
            'Английский язык': [''] * 5,
            'Информатика': [''] * 5
        }
        df_9b = pd.DataFrame(subjects_9b)
        df_9b.to_excel(writer, sheet_name='9Б', index=False)
        
        # Класс 10А - с профильными предметами
        subjects_10a = {
            'Алгебра': [''] * 5,
            'Геометрия': [''] * 5,
            'Русский язык': [''] * 5,
            'Литература': [''] * 5,
            'Физика': [''] * 5,
            'Химия': [''] * 5,
            'Биология': [''] * 5,
            'История': [''] * 5,
            'Обществознание': [''] * 5,
            'Английский язык': [''] * 5,
            'Информатика': [''] * 5,
            'Экономика': [''] * 5
        }
        df_10a = pd.DataFrame(subjects_10a)
        df_10a.to_excel(writer, sheet_name='10А', index=False)
        
        # Класс 11В
        subjects_11v = {
            'Алгебра': [''] * 5,
            'Геометрия': [''] * 5,
            'Русский язык': [''] * 5,
            'Литература': [''] * 5,
            'Физика': [''] * 5,
            'Химия': [''] * 5,
            'Биология': [''] * 5,
            'История': [''] * 5,
            'Обществознание': [''] * 5,
            'Английский язык': [''] * 5
        }
        df_11v = pd.DataFrame(subjects_11v)
        df_11v.to_excel(writer, sheet_name='11В', index=False)
    
    # Форматирование
    if HAS_OPENPYXL:
        format_classes_excel(filename)
    
    print(f"✅ Создан файл: {filename}")
    print(f"   📋 Листы: 9А, 9Б, 10А, 11В")
    return filename


def format_excel(filename, sheet_name):
    """Форматирует Excel файл (заголовки)"""
    try:
        wb = load_workbook(filename)
        ws = wb[sheet_name]
        
        # Форматирование заголовков
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Автоширина столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
    except Exception as e:
        print(f"Ошибка форматирования: {e}")


def format_classes_excel(filename):
    """Форматирует Excel файл с классами"""
    try:
        wb = load_workbook(filename)
        
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Форматирование заголовков
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Автоширина
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
    except Exception as e:
        print(f"⚠️ Ошибка форматирования: {e}")


def main():
    """Главная функция"""
    print("=" * 60)
    print("🎓 ГЕНЕРАТОР ПРИМЕРОВ EXCEL ФАЙЛОВ ДЛЯ EDUCATION PLATFORM")
    print("=" * 60)
    print()
    
    # Создаем файлы
    print("📝 Создание примеров файлов...")
    print()
    
    students_file = create_students_example()
    teachers_file = create_teachers_example()
    classes_file = create_classes_subjects_example()
    
    print()
    print("=" * 60)
    print("✨ ВСЕ ФАЙЛЫ УСПЕШНО СОЗДАНЫ!")
    print("=" * 60)
    print()
    print("📁 Созданные файлы:")
    print(f"   1️⃣  {students_file} - Пример учеников (10 записей)")
    print(f"   2️⃣  {teachers_file} - Пример учителей (8 записей)")
    print(f"   3️⃣  {classes_file} - Пример классов и предметов (4 класса)")
    print()
    print("📖 Откройте файл ИНСТРУКЦИЯ.md для подробного руководства")
    print()
    print("🚀 Используйте эти файлы как шаблоны для загрузки данных!")
    print("=" * 60)


if __name__ == "__main__":
    main()